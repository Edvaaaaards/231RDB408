import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook

# Inicializē Selenium un atver pārlūku
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

# Iegūst darbinieku pilnos vārdus no people.csv faila
name = []
with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row = line.rstrip().split(",") 
        name.append(row[0])  # Izmanto tikai pirmo elementu

# Iegūst informāciju par algām no salary.xlsx faila
salaries = {}
workbook = load_workbook("salary.xlsx")
sheet = workbook.active
for row in sheet.iter_rows(min_row=2, values_only=True):
    full_name, salary = row
    salaries[full_name] = salary

# Aizver salary.xlsx failu, jo vairs nav nepieciešams
workbook.close()

# Aizpilda CRC32 vērtības un izdrukā rezultātus
url = "https://emn178.github.io/online-tools/crc32.html"
for full_name in name:
    driver.get(url)
    time.sleep(2)

    find = driver.find_element(By.ID, "input")
    find.clear()
    find.send_keys(full_name)

    find = driver.find_element(By.ID, "output")
    crc32_result = find.get_attribute("value")

    print(f"Full Name: {full_name}, CRC32: {crc32_result}, Salary: {salaries.get(full_name, 'N/A')}")

# Aizver pārlūku
driver.quit()
