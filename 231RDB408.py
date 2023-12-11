import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from openpyxl import load_workbook

service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

vards = []
name = []

with open("people.csv", "r") as file:
    next(file)
    for line in file:
        row = line.rstrip().split(",")
        name.append(row)

for i in name:
    b = i[2] + " " + i[3]
    vards.append(b)

url = "https://emn178.github.io/online-tools/crc32.html"
driver.get(url)
time.sleep(2)

salaries = {}
workbook = load_workbook("salary.xlsx")
sheet = workbook['Sheet2']
max_row = sheet.max_row

for full_name in vards:
    find = driver.find_element(By.ID, "input")
    find.clear()
    find.send_keys(full_name)

    find = driver.find_element(By.ID, "output")
    crc32_result = find.get_attribute("value")

    if crc32_result in salaries:
        salaries[crc32_result] += otrais
    else:
        total_salary = 0
        for row in range(2, max_row + 1):
            pirmais = sheet['A' + str(row)].value
            otrais = sheet['B' + str(row)].value

            if pirmais == crc32_result:
                total_salary += otrais

        salaries[crc32_result] = total_salary

    print(f"pilns vārds: {full_name}, CRC32: {crc32_result}, Total Salary: {salaries.get(crc32_result, 'N/A')}")     

workbook.close()
driver.quit()
